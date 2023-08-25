from datetime import datetime
from openpyxl import load_workbook
import codecs

# Author:   w26827, Michael. Nielsen@ufst.dk
# Last edit:    2022-06-24

class CprDemoTool:
    # Text

    # Job Counters:
    cnt_Set_Rec = 0  # Counts the number of records in the current Set
    cnt_Tot_Set = 0  # Counts the number of Sets, which have been generated!
    cnt_Tot_Rec = 0  # Counts the total number of records, which have been generated!
    # Record Counters:
    cnt_Rec000 = 0
    cnt_Rec001 = 0
    cnt_Rec002 = 0
    cnt_Rec003 = 0
    cnt_Rec004 = 0
    cnt_Rec008 = 0
    cnt_Rec009 = 0
    cnt_Rec010 = 0
    cnt_Rec017 = 0
    cnt_Rec052 = 0
    cnt_Rec099 = 0
    cnt_Rec999 = 0

    # Record Lenghts:
    # len000 = 35
    # len001 = 106
    # len002 = 306
    # len003 = 249
    # len004 = 37
    # len008 = 193
    # len009 = 37
    # len010 = 30
    # len017 = 272
    # len052 = 287
    # len099 = 45
    # len999 = 21

    # body of constructor
    def __init__(self):
        self.a10_PNR = ' ' * 10
        self.a10_SORTFELT_10 = '0000000000'
        self.a10_SORTFELT_10 = '9999999999'
        self.n06_OPGAVENR = '999999'
        self.a08_PRODDTO = 'YYYYMMDD'
        self.a08_PRODDTOFORRIG = 'YYYYMMDD'
        self.a10_PNRGAELD = ' ' * 10
        self.n02_STATUS = '01'
        self.n12_STATUSHAENSTART = '000000000000'
        self.a01_STATUSDTO_UMRK = ' '
        self.a01_KOEN = ' '
        self.a10_FOED_DT = ' ' * 10
        self.a01_FOED_DT_UMRK = ' '
        self.a10_START_DTPERSON = ' ' * 10
        self.a01_START_DT_UMRKPERSON = ' '
        self.a10_SLUT_DT_PERSON = ' ' * 10
        self.a01_SLUT_DT_UMRKPERSON = ' '
        self.a34_STILLING = "                                   "
        self.n04_KOMKOD = '9999'
        self.n04_VEJKOD = '9999'
        self.a04_HUSNR = '    '
        self.a02_ETAGE = '  '
        self.a04_SIDEDOER = '    '
        self.a04_BNR = '    '
        self.a34_CONVN = '                                  '
        self.a34_LOKALITET = ' ' * 34
        self.a34_STANDARDADR = ' ' * 34
        self.a20_VEJADRNVN = ' ' * 20
        self.a34_BYNAVN = ' ' * 34
        self.a04_POSTNR = ' ' * 4
        self.a20_POSTDISTTXT = ' ' * 20
        self.n12_TILFLYDTO = '000000000000'
        self.a01_TILFLYDTO_UMRK = ' '
        self.n12_TILFLYKOMDTO = '000000000000'
        self.a01_TILFLYKOMDT_UMRK = ' '
        self.n04_FRAFLYKOMKOD = '9999'
        self.n12_FRAFLYKOMDTO = '000000000000'
        self.a01_FRAFLYKOMDT_UMRK = ' '
        self.n04_START_MYNKODADRTXT = '9999'
        self.a34_ADR1_SUPLADR = '                                  '
        self.a34_ADR2_SUPLADR = '                                  '
        self.a34_ADR3_SUPLADR = '                                  '
        self.a34_ADR4_SUPLADR = '                                  '
        self.a34_ADR5_SUPLADR = '                                  '
        self.a10_START_DTADRTXT = '          '
        self.a10_SLET_DT_ADRTXT = '          '
        self.n04_BESKYTTYPE = '0001'
        self.a10_START_DTBESKYTTELSE = 'ÅÅÅÅ-MM-DD'
        self.a10_SLET_DTBESKYTTELSE = 'ÅÅÅÅ-MM-DD'
        self.a50_FORNVN = '                                                  '
        self.a01_FORNVN_MRK = ' '
        self.a40_MELNVN = '                                        '
        self.a01_MELNVN_MRK = ' '
        self.a40_EFTERNVN = '                                        '
        self.a01_EFTERNVN_MRK = ' '
        self.n12_NVNHAENSTART = '000000000000'
        self.a01_HAENSTART_UMRK_NAVNE = ' '
        self.a34_ADRNVN = '                                  '
        self.n04_START_MYNKODFODESTED = '0000'
        self.a20_MYNTXTFODESTED = '                    '
        self.n04_LANDEKODE = '0000'
        self.n12_HAENSTARTSTATSBORGERSKAB = '000000000000'
        self.a01_HAENSTART_UMRKSTATSBORGERSKAB = ' '
        self.a10_START_DTUMYNDIG = ' ' * 12
        self.a01_START_DT_UMRKUMYNDIG = ' '
        self.a10_SLET_DTUMYNDIG = ' ' * 10
        self.n04_UMYN_RELTYP = "0001"  # 1 = værge PNR findes
        self.n10_RELPNR = ' ' * 10
        self.a10_START_DTRELPNR_PNR = ' ' * 10
        self.a15_VAERGEMAALS_TYPE = ' ' * 15
        self.a170_FILLER = ' ' * 170
        self.a214_FILLER = ' ' * 214
        self.a224_FILLER = ' ' * 224
        self.a12_TIMESTAMPU = 'ÅÅÅÅMMDDTTMM'
        self.a03_HAENDELSE = '   '
        self.a02_AFLEDTMRK = '  '
        self.a15_NGLKONST = '               '

    def checknumberformat(self, number, ciff, oper):
        lnum = len(number)
        if not number.isnumeric():
            print("Feltet må kun bestå af ciffre - prøv igen!")
            return 1
        else:
            if oper == "ge" and lnum < ciff:
                print("Feltet skal bestå af mindst 6 ciffre - prøv igen!")
                return 2
            if oper == "le" and lnum > ciff:
                print("Feltet skal bestå af højest 6 ciffre - prøv igen!")
                return 3
            if oper == "eq" and lnum != ciff:
                print("Feltet skal bestå af 6 ciffre - prøv igen!")
                return 4
            if oper == "ne" and lnum == ciff:
                print("Feltet må ikke bestå af 6 ciffre - prøv igen!")
                return 5
        return 0

    def checkdateformat(self, dato):
        # Verificerer at datoerne har det rette format
        # Return 3 if the date doesn't contain 8 digits
        # Return 2 if the date is not numeric
        # Return 1 if the date is wrong
        # Return 0 if the date is in the correct format

        if len(dato) != 8:
            print("Taste FEJL: Der skal være 8 ciffre")
            return 3
        if not dato.isdigit():
            print("Taste FEJL: Der må kun være ciffre")
            return 2
        if int(dato[:4]) < 1900 or int(dato[:4]) > 2500:
            print("Taste FEJL: Årstallet er forkert: ", dato[:4])
            return 1
        if int(dato[4:6]) < 1 or int(dato[4:6]) > 12:
            print("Taste FEJL: Måneden er forkert: ", dato[4:6])
            return 1
        if int(dato[6:8]) < 1 or int(dato[6:8]) > 31:
            print("Taste FEJL: Dagen er forkert: ", dato[6:8])
            return 1
        return 0

    # Strips dates down to YYYY-MM-DD form
    def fdto(self, dto):
        if (len(str(dto).strip())) == 0 or dto is None:
            return ' ' * 10
        else:
            return str(dto)[:10]

    # Removes 'None' from empty fields in the Spreadsheet
    def fstr(self, field, spaces):
        if field is None:
            return ' ' * spaces
        else:
            return field

    # Indlæser Kørsels Data
    def get_task_data(self):

        # Indlæser Opgavenummer
        self.n06_OPGAVENR = input("Indtast 6 Cifferet OpgaveNummer: ")
        while self.checknumberformat(self.n06_OPGAVENR, 6, "eq") > 0:
            self.n06_OPGAVENR = input("Indtast 6 Cifferet OpgaveNummer: ")

        # Indlæser Produktionsdato (ÅÅÅÅMMDD) og Forrige produktionsdato ÅÅÅÅMMDD
        while True:
            self.a08_PRODDTO = input("Indtast Produktionsdato på formen ""ÅÅÅÅMMDD"": ")
            while cprDemoTool.checkdateformat(self.a08_PRODDTO) > 0:
                self.a08_PRODDTO = input("Indtast Produktionsdato på formen ""ÅÅÅÅMMDD"": ")

            self.a08_PRODDTOFORRIG = input("Indtast Forrige Produktionsdato på formen ""ÅÅÅÅMMDD"": ")
            while cprDemoTool.checkdateformat(self.a08_PRODDTOFORRIG) > 0:
                self.a08_PRODDTOFORRIG = input("Indtast Forrige Produktionsdato på formen ""ÅÅÅÅMMDD"": ")

            if self.a08_PRODDTO > self.a08_PRODDTOFORRIG:
                break
            elif self.a08_PRODDTO == self.a08_PRODDTOFORRIG:
                print("Taste FEJL: Produktionsdatoen og Forrige Produktionsdato ligger indenfor samme døgn!")

            else:
                print("Taste FEJL: Produktionsdatoen ligger før Forrige Produktionsdato")

    # 000 Startrecord
    def make_rec_000(self):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('000')

        # Skriver SORTFELT-10
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_SORTFELT_10 = '0000000000'
        fOut.write(f'{self.a10_SORTFELT_10:>10}')

        # Skriver OPGAVENR, Opgavenummer
        if not fOut.tell() == 13:
            fOut.seek(13)
        fOut.write(str(self.n06_OPGAVENR))

        # Skriver PRODDTO, Produktionsdato (ÅÅÅÅMMDD)
        if not fOut.tell() == 19:
            fOut.seek(19)
        fOut.write(self.a08_PRODDTO)

        # Skriver PRODDTOFORRIG, Forrige produktionsdato ÅÅÅÅMMDD
        if not fOut.tell() == 27:
            fOut.seek(27)
        fOut.write(self.a08_PRODDTOFORRIG + '\n')

        self.cnt_Rec000 += 1
        self.cnt_Set_Rec = 0

        return 0

    # 001 Personoplysninger
    def make_rec_001(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('001')

        # Skriver PNR, Personnummer
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = cprDemoTool.fstr(wbs['A' + str(row)].value, 10)  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Skriver PNRGAELD, Gældende personnummer
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.a10_PNRGAELD = cprDemoTool.fstr(wbs['B' + str(row)].value, 10)
        fOut.write(f'{str(self.a10_PNRGAELD):>010}')

        # Statuskoder:
        # 01 = Aktiv, bopæl i dansk folkeregister
        # 03 = Aktiv, speciel vejkode (9900 - 9999) i dansk folke register
        # 05 = Aktiv, bopæl i grønlandsk folkeregister
        # 07 = Aktiv, speciel vejkode (9900 - 9999) i grønlandsk folkeregister
        # 20 = Inaktiv, uden bopæl i dansk/grønlandsk folkeregister
        #     men tildelt personnummer af skattehensyn (kommunekoderne 0010, 0011, 0012 og 0019)
        # 30 = Inaktiv, annulleret personnummer
        # 50 = Inaktiv, slettet personnummer ved dobbeltnummer
        # 60 = Inaktiv, ændret personnummer ved ændring af fødselsdato og køn
        # 70 = Inaktiv, forsvundet
        # 80 = Inaktiv, udrejst
        # 90 = Inaktiv, død

        # Skriver STATUS, Status Kode
        if not fOut.tell() == 23:
            fOut.seek(23)
        self.n02_STATUS = cprDemoTool.fstr(wbs['C' + str(row)].value, 2)
        fOut.write(f'{str(self.n02_STATUS):>2}')

        # Skriver STATUSHAENSTART, Statusdato
        if not fOut.tell() == 25:
            fOut.seek(25)
        fOut.write(f'{str(self.n12_STATUSHAENSTART):>12}')

        # Skriver STATUSDTO_UMRK, Statusdato usikkerhedsmarkering
        if not fOut.tell() == 37:
            fOut.seek(37)
        fOut.write(f'{str(self.a01_STATUSDTO_UMRK):>1}')

        # Skriver KOEN, Køn Værdisæt: M = mænd K = kvinder
        if not fOut.tell() == 38:
            fOut.seek(38)
        self.a01_KOEN = cprDemoTool.fstr(wbs['D' + str(row)].value, 1)
        fOut.write(f'{str(self.a01_KOEN):>1}')

        # Skriver FOED_DT, Fødselsdato ÅÅÅÅ-MM-DD
        if not fOut.tell() == 39:
            fOut.seek(39)
        self.a10_FOED_DT = cprDemoTool.fdto(wbs['E' + str(row)].value)
        fOut.write(self.a10_FOED_DT)

        # Skriver FOED_DT_UMRK, Fødselsdato usikkerhedsmarkering
        if not fOut.tell() == 49:
            fOut.seek(49)
        fOut.write(f'{str(self.a01_FOED_DT_UMRK):>1}')

        # Skriver START_DTPERSON, Person startdato ÅÅÅÅ-MM-DD
        if not fOut.tell() == 50:
            fOut.seek(50)
        self.a10_START_DTPERSON = cprDemoTool.fdto(wbs['F' + str(row)].value)
        fOut.write(self.a10_START_DTPERSON)

        # Skriver START_DT_UMRKPERSON, Startdato usikkerhedsmarkering
        if not fOut.tell() == 60:
            fOut.seek(60)
        fOut.write(f'{str(self.a01_START_DT_UMRKPERSON):>1}')

        # Skriver SLUT_DTPERSON, Person slutdato ÅÅÅÅ-MM-DD
        if not fOut.tell() == 61:
            fOut.seek(61)
        fOut.write(self.a10_SLUT_DT_PERSON)

        # Skriver SLUT_DT_UMRKPERSON, Slutdato usikkerhedsmarkering
        if not fOut.tell() == 71:
            fOut.seek(71)
        fOut.write(f'{str(self.a01_SLUT_DT_UMRKPERSON):>1}')

        # Skriver STILLING, Stilling
        if not fOut.tell() == 72:
            fOut.seek(72)
        self.a34_STILLING = cprDemoTool.fstr(wbs['G' + str(row)].value, 34)
        fOut.write(f'{str(self.a34_STILLING):<34}' + '\n')

        self.cnt_Rec001 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 002 Aktuelle adresseoplysninger
    def make_rec_002(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('002')

        # Skriver PNR, Personnummer
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Skriver KOMKOD, kommunenummer
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.n04_KOMKOD = cprDemoTool.fstr(wbs['H' + str(row)].value, 4)  # get Kommunekode
        fOut.write(f'{str(self.n04_KOMKOD):>4}')

        # Skriver VEJKOD, vejkode
        if not fOut.tell() == 17:
            fOut.seek(17)
        self.n04_VEJKOD = cprDemoTool.fstr(wbs['I' + str(row)].value, 4)  # get vejkode
        fOut.write(f'{str(self.n04_VEJKOD):>4}')

        # Skriver HUSNR, husnummer
        if not fOut.tell() == 21:
            fOut.seek(21)
        self.a04_HUSNR = cprDemoTool.fstr(wbs['J' + str(row)].value, 4)  # get husnummer
        fOut.write(f'{str(self.a04_HUSNR):>4}')

        # Skriver ETAGE, etage
        if not fOut.tell() == 25:
            fOut.seek(25)
        self.a02_ETAGE = cprDemoTool.fstr(wbs['K' + str(row)].value, 2)  # get etage
        fOut.write(f'{str(self.a02_ETAGE):>2}')

        # Skriver SIDEDOER, Sidedør nummer
        if not fOut.tell() == 27:
            fOut.seek(27)
        self.a04_SIDEDOER = cprDemoTool.fstr(wbs['L' + str(row)].value, 4)  # get sidedør
        fOut.write(f'{str(self.a04_SIDEDOER):>4}')

        # Skriver BNR, Bygningsnummer
        if not fOut.tell() == 31:
            fOut.seek(31)
        self.a04_BNR = cprDemoTool.fstr(wbs['M' + str(row)].value, 4)  # get Bygningsnummer
        fOut.write(f'{str(self.a04_BNR):>4}')

        # Skriver CONVN, C/O navn
        if not fOut.tell() == 35:
            fOut.seek(35)
        self.a34_CONVN = cprDemoTool.fstr(wbs['N' + str(row)].value, 34)  # get C/O navn
        fOut.write(f'{str(self.a34_CONVN):<34}')

        # Skriver TILFLYDTO, Tilflytningsdato
        if not fOut.tell() == 69:
            fOut.seek(69)
        fOut.write(f'{str(self.n12_TILFLYDTO):>12}')

        # Skriver a01_TILFLYDTO_UMRK,Tilflytningsdato usikkerhedsmarkering
        if not fOut.tell() == 81:
            fOut.seek(81)
        fOut.write(f'{str(self.a01_TILFLYDTO_UMRK):>1}')

        # Skriver TILFLYKOMDTO, Tilflytning kommune dato
        if not fOut.tell() == 82:
            fOut.seek(82)
        fOut.write(f'{str(self.n12_TILFLYDTO):>12}')

        # Skriver TILFLYKOMDT_UMRK, Tilflytning kommune dato usikkerhedsmarkering
        if not fOut.tell() == 94:
            fOut.seek(94)
        fOut.write(f'{str(self.a01_TILFLYKOMDT_UMRK):>1}')

        # Skriver n04_FRAFLYKOMKOD, Fraflytning kommunekode
        if not fOut.tell() == 95:
            fOut.seek(95)
        self.n04_FRAFLYKOMKOD = cprDemoTool.fstr(wbs['O' + str(row)].value, 4)  # get C/O navn
        fOut.write(f'{str(self.n04_FRAFLYKOMKOD):>4}')

        # Skriver FRAFLYKOMDTO, Fraflytning kommune dato
        if not fOut.tell() == 99:
            fOut.seek(99)
        fOut.write(f'{str(self.n12_FRAFLYKOMDTO):>12}')

        # Skriver FRAFLYKOMDT_UMRK, Fraflytning kommune dato usikkerhedsmarkering
        if not fOut.tell() == 111:
            fOut.seek(111)
        fOut.write(f'{str(self.a01_FRAFLYKOMDT_UMRK):>1}')

        # Skriver START_MYNKODADRTXT, Start myndighed
        if not fOut.tell() == 112:
            fOut.seek(112)
        fOut.write(f'{str(self.n04_START_MYNKODADRTXT):>4}')

        # Skriver ADR1-SUPLADR, 1. linie af supplerende adr
        if not fOut.tell() == 116:
            fOut.seek(116)
        fOut.write(f'{str(self.a34_ADR1_SUPLADR):<34}')

        # Skriver ADR2-SUPLADR, 2. linie af supplerende adr
        if not fOut.tell() == 150:
            fOut.seek(150)
        fOut.write(f'{str(self.a34_ADR2_SUPLADR):<34}')

        # Skriver ADR3-SUPLADR, 3. linie af supplerende adr
        if not fOut.tell() == 184:
            fOut.seek(184)
        fOut.write(f'{str(self.a34_ADR3_SUPLADR):<34}')

        # Skriver ADR4-SUPLADR, 4. linie af supplerende adr
        if not fOut.tell() == 218:
            fOut.seek(218)
        fOut.write(f'{str(self.a34_ADR4_SUPLADR):<34}')

        # Skriver ADR5-SUPLADR, 5. linie af supplerende adr
        if not fOut.tell() == 252:
            fOut.seek(252)
        fOut.write(f'{str(self.a34_ADR5_SUPLADR):<34}')

        # Skriver START_DTADRTXT, Startdato ÅÅÅÅ-MM-DD
        if not fOut.tell() == 286:
            fOut.seek(286)
        fOut.write(f'{str(self.a10_START_DTADRTXT):>10}')

        # Skriver SLET_DT_ADRTXT, Slettedato ÅÅÅÅ-MM-DD
        if not fOut.tell() == 296:
            fOut.seek(296)
        fOut.write(f'{str(self.a10_SLET_DT_ADRTXT):>10}' + '\n')

        self.cnt_Rec002 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 003 Adrnvn og klarskriftadresse
    def make_rec_003(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('003')

        # Skriver PNR, Personnummer
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Skriver ADRNVN, Adresseringsnavn
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.a34_ADRNVN = cprDemoTool.fstr(wbs['AA' + str(row)].value, 34)  # get Adresseringsnavn
        fOut.write(f'{str(self.a34_ADRNVN):<34}')

        # Skriver CONVN, C/O navn
        if not fOut.tell() == 47:
            fOut.seek(47)
        self.a34_CONVN = cprDemoTool.fstr(wbs['N' + str(row)].value, 34)  # get C/O navn
        fOut.write(f'{str(self.a34_CONVN):<34}')

        # Skriver LOKALITET, Lokalitet
        if not fOut.tell() == 81:
            fOut.seek(81)
        self.a34_LOKALITET = cprDemoTool.fstr(wbs['AB' + str(row)].value, 34)  # get Lokalitet
        fOut.write(f'{str(self.a34_LOKALITET):<34}')

        # Skriver STANDARDADR, Vejadrnvn,husnr,etage,sidedoer, bnr., Etiketteadresse
        if not fOut.tell() == 115:
            fOut.seek(115)
        self.a34_STANDARDADR = cprDemoTool.fstr(wbs['P' + str(row)].value, 34)  # get Etiketteadresse
        fOut.write(f'{str(self.a34_STANDARDADR):<34}')

        # Skriver BYNAVN, Bynavn
        if not fOut.tell() == 149:
            fOut.seek(149)
        self.a34_BYNAVN = cprDemoTool.fstr(wbs['R' + str(row)].value, 34)  # get Bynavn
        fOut.write(f'{str(self.a34_BYNAVN):<34}')

        # Skriver POSTNR, Postnummer
        if not fOut.tell() == 183:
            fOut.seek(183)
        self.a04_POSTNR = cprDemoTool.fstr(wbs['S' + str(row)].value, 4)  # get Postnummer
        fOut.write(f'{str(self.a04_POSTNR):<4}')

        # Skriver POSTDISTTXT, PostDistrikt
        if not fOut.tell() == 187:
            fOut.seek(187)
        self.a20_POSTDISTTXT = cprDemoTool.fstr(wbs['T' + str(row)].value, 20)  # get Postdistrikt tekst
        fOut.write(f'{str(self.a20_POSTDISTTXT):<20}')

        # Skriver KOMKOD, kommunenummer
        if not fOut.tell() == 207:
            fOut.seek(207)
        self.n04_KOMKOD = cprDemoTool.fstr(wbs['H' + str(row)].value, 4)  # get Kommunekode
        fOut.write(f'{str(self.n04_KOMKOD):>4}')

        # Skriver VEJKOD, vejkode
        if not fOut.tell() == 211:
            fOut.seek(211)
        self.n04_VEJKOD = cprDemoTool.fstr(wbs['I' + str(row)].value, 4)  # get vejkode
        fOut.write(f'{str(self.n04_VEJKOD):>4}')

        # Skriver HUSNR, husnummer
        if not fOut.tell() == 215:
            fOut.seek(215)
        self.a04_HUSNR = cprDemoTool.fstr(wbs['J' + str(row)].value, 4)  # get husnummer
        fOut.write(f'{str(self.a04_HUSNR):>4}')

        # Skriver ETAGE, etage
        if not fOut.tell() == 219:
            fOut.seek(219)
        self.a02_ETAGE = cprDemoTool.fstr(wbs['K' + str(row)].value, 2)  # get etage
        fOut.write(f'{str(self.a02_ETAGE):>2}')

        # Skriver SIDEDOER, Sidedør nummer
        if not fOut.tell() == 221:
            fOut.seek(221)
        self.a04_SIDEDOER = cprDemoTool.fstr(wbs['L' + str(row)].value, 4)  # get sidedør
        fOut.write(f'{str(self.a04_SIDEDOER):>4}')

        # Skriver BNR, Bygningsnummer
        if not fOut.tell() == 225:
            fOut.seek(225)
        self.a04_BNR = cprDemoTool.fstr(wbs['M' + str(row)].value, 4)  # get Bygningsnummer
        fOut.write(f'{str(self.a04_BNR):>4}')

        # Skriver VEJADRNVN, Vejadresseringsnavn
        if not fOut.tell() == 229:
            fOut.seek(229)
        self.a20_VEJADRNVN = cprDemoTool.fstr(wbs['Q' + str(row)].value, 20)  # get Vejadresseringsnavn
        fOut.write(f'{str(self.a20_VEJADRNVN):<20}' + '\n')

        self.cnt_Rec003 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 004 Beskyttelse
    def make_rec_004(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('004')

        # Skriver PNR
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Beskyttelsestyper
        # 0001 = navne- og adressebeskyttelse
        # 0002 = lokalvejviserbeskyttelse
        # 0003 = reklamebeskyttelse
        # 0004 = forskerbeskyttelse

        # Skriver BESKYTTYPE
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.n04_BESKYTTYPE = cprDemoTool.fstr(wbs['U' + str(row)].value, 4)  # 0001 = navne- og adressebeskyttelse
        fOut.write(f'{str(self.n04_BESKYTTYPE):>4}')

        # Indlæser og skriver START_DTBESKYTTELSE (Dato for ikrafttræden af databeskyttelse)
        if not fOut.tell() == 17:
            fOut.seek(17)
        self.a10_START_DTBESKYTTELSE = cprDemoTool.fdto(wbs['V' + str(row)].value)
        fOut.write(self.a10_START_DTBESKYTTELSE)

        # Indlæser og skriver SLET_DTBESKYTTELSE (Dato for ophør af databeskyttelse)
        if not fOut.tell() == 27:
            fOut.seek(27)
        self.a10_SLET_DTBESKYTTELSE = cprDemoTool.fdto(wbs['W' + str(row)].value)
        fOut.write(self.a10_SLET_DTBESKYTTELSE + '\n')

        self.cnt_Rec004 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 008 Aktuelle navneoplysninger
    def make_rec_008(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('008')

        # Skriver PNR
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Skriver FORNVN, Fornavn(e)
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.a50_FORNVN = cprDemoTool.fstr(wbs['X' + str(row)].value, 50)  # get CPR Nummer
        fOut.write(f'{str(self.a50_FORNVN):<50}')

        # navne Markeringer:
        # + navn er forkortet
        # * navn indeholder tegn, der ikke kan indrapporteres til CPR.
        # = navn er ikke dokumenteret

        # Skriver FORNVN_MRK, Fornavn markering
        if not fOut.tell() == 63:
            fOut.seek(63)
        fOut.write(f'{str(self.a01_FORNVN_MRK):>1}')

        # Skriver MELNVN, Mellemnavn
        if not fOut.tell() == 64:
            fOut.seek(64)
        self.a40_MELNVN = cprDemoTool.fstr(wbs['Y' + str(row)].value, 40)
        fOut.write(f'{str(self.a40_MELNVN):<40}')

        # Skriver MELNVN_MRK, Mellemnavn markering
        if not fOut.tell() == 104:
            fOut.seek(104)
        fOut.write(f'{str(self.a01_MELNVN_MRK):>1}')

        # Skriver EFTERNVN, Efternavn
        if not fOut.tell() == 105:
            fOut.seek(105)
        self.a40_EFTERNVN = cprDemoTool.fstr(wbs['Z' + str(row)].value, 40)
        fOut.write(f'{str(self.a40_EFTERNVN):<40}')

        # Skriver EFTERNVN_MRK, Efternavn markering
        if not fOut.tell() == 145:
            fOut.seek(145)
        fOut.write(f'{str(self.a01_EFTERNVN_MRK):>1}')

        # Skriver NVNHAENSTART, Navne startdato
        if not fOut.tell() == 146:
            fOut.seek(146)
        fOut.write(f'{str(self.n12_NVNHAENSTART):>12}')

        # Skriver HAENSTART_UMRK-NAVNE, Navne startdato usikkerhedsmarkering
        if not fOut.tell() == 158:
            fOut.seek(158)
        fOut.write(f'{str(self.a01_HAENSTART_UMRK_NAVNE):>1}')

        # Skriver ADRNVN, Adresseringsnavn
        if not fOut.tell() == 159:
            fOut.seek(159)
        self.a34_ADRNVN = cprDemoTool.fstr(wbs['AA' + str(row)].value, 34)
        fOut.write(f'{str(self.a34_ADRNVN):<34}' + '\n')

        self.cnt_Rec008 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 009 Fødselsregistreringsoplysninger
    def make_rec_009(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('009')

        # Skriver PNR
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Skriver START_MYNKODFØDESTED, Fødselsregistreringssteds kode
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.n04_START_MYNKODFODESTED = cprDemoTool.fstr(wbs['AN' + str(row)].value, 4)
        fOut.write(f'{str(self.n04_START_MYNKODFODESTED):>4}')

        # Skriver MYNTXTFØDESTED, Supplerende fødselsregistreringssted tekst
        if not fOut.tell() == 17:
            fOut.seek(17)
        fOut.write(f'{str(self.a20_MYNTXTFODESTED):<20}' + '\n')

        self.cnt_Rec009 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 010 Aktuelt statsborgerskab
    def make_rec_010(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('010')

        # Skriver PNR
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Skriver LANDEKODE,
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.n04_LANDEKODE = cprDemoTool.fstr(wbs['AC' + str(row)].value, 4)
        fOut.write(f'{str(self.n04_LANDEKODE):>4}')

        # Skriver n12_HAENSTARTSTATSBORGERSKAB, Statsborgerskab startdato
        if not fOut.tell() == 17:
            fOut.seek(17)
        fOut.write(f'{str(self.n12_HAENSTARTSTATSBORGERSKAB):>12}')

        # Skriver HAENSTART_UMRKSTATSBORGERSKAB,
        if not fOut.tell() == 29:
            fOut.seek(29)
        fOut.write(f'{str(self.a01_HAENSTART_UMRKSTATSBORGERSKAB):>1}' + '\n')

        self.cnt_Rec010 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 017 Umyndiggørelse og værge
    def make_rec_017(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('017')

        # Skriver PNR
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Indlæser og skriver START_DTUMYNDIG, Umyndiggørelse startdato ÅÅÅÅ-MM-DD
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.a10_START_DTUMYNDIG = cprDemoTool.fdto(wbs['AD' + str(row)].value)
        fOut.write(self.a10_START_DTUMYNDIG)

        # Skriver START_DT_UMRKUMYNDIG, Startdato usikkerhedsmarkering
        if not fOut.tell() == 23:
            fOut.seek(23)
        fOut.write(f'{str(self.a01_START_DT_UMRKUMYNDIG):>1}')

        # Indlæser og skriver SLET_DTUMYNDIG  (Umyndiggørelse slutdato ÅÅÅÅ-MM-DD)
        if not fOut.tell() == 24:
            fOut.seek(24)
        self.a10_SLET_DTUMYNDIG = cprDemoTool.fdto(wbs['AE' + str(row)].value)
        fOut.write(self.a10_SLET_DTUMYNDIG)

        # Under-værgemål relationstype:
        # 1 = værge er indsat med personnummer
        # 2 = værge findes med navn og adresse

        # Indlæser Skriver UMYN_RELTYP, Under værgemål relationstype
        if not fOut.tell() == 34:
            fOut.seek(34)
        self.n04_UMYN_RELTYP = cprDemoTool.fstr(wbs['AF' + str(row)].value, 4)
        fOut.write(f'{str(self.n04_UMYN_RELTYP):>04}')

        # Indlæser og skriver RELPNR, Relation personnummer
        if not fOut.tell() == 38:
            fOut.seek(38)
        self.n10_RELPNR = cprDemoTool.fstr(wbs['AG' + str(row)].value, 10)
        fOut.write(f'{str(self.n10_RELPNR):0>10}')

        # Skriver FILLER til EOF (214 spaces)
        if not fOut.tell() == 58:
            fOut.seek(58)
        fOut.write(self.a214_FILLER + '\n')

        self.cnt_Rec017 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 052 Umyndiggørelse og værge med værgemåltype
    def make_rec_052(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('052')

        # Skriver PNR
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Indlæser og skriver START_DTUMYNDIG, Umyndiggørelse startdato ÅÅÅÅ-MM-DD
        if not fOut.tell() == 13:
            fOut.seek(13)
        self.a10_START_DTUMYNDIG = cprDemoTool.fdto(wbs['AD' + str(row)].value)
        fOut.write(self.a10_START_DTUMYNDIG)

        # Skriver START_DT_UMRKUMYNDIG, Startdato usikkerhedsmarkering
        if not fOut.tell() == 23:
            fOut.seek(23)
        fOut.write(f'{str(self.a01_START_DT_UMRKUMYNDIG):>1}')

        # Indlæser og skriver SLET_DTUMYNDIG  (Umyndiggørelse slutdato ÅÅÅÅ-MM-DD)
        if not fOut.tell() == 24:
            fOut.seek(24)
        self.a10_SLET_DTUMYNDIG = cprDemoTool.fdto(wbs['AE' + str(row)].value)
        fOut.write(self.a10_SLET_DTUMYNDIG)

        # Under-værgemål relationstype:
        # 1 = værge er indsat med personnummer
        # 2 = værge findes med navn og adresse

        # Indlæser Skriver UMYN_RELTYP, Under værgemål relationstype
        if not fOut.tell() == 34:
            fOut.seek(34)
        self.n04_UMYN_RELTYP = cprDemoTool.fstr(wbs['AF' + str(row)].value, 4)
        fOut.write(f'{str(self.n04_UMYN_RELTYP):>04}')

        # Indlæser og skriver RELPNR, Relation personnummer
        if not fOut.tell() == 38:
            fOut.seek(38)
        self.n10_RELPNR = cprDemoTool.fstr(wbs['AG' + str(row)].value, 10)
        fOut.write(f'{str(self.n10_RELPNR):0>10}')

        # Indlæser og skriver START_DTRELPNR_PNR, Relation personnummer startdato
        if not fOut.tell() == 48:
            fOut.seek(48)
        self.a10_START_DTRELPNR_PNR = cprDemoTool.fdto(wbs['AH' + str(row)].value)
        fOut.write(self.a10_START_DTRELPNR_PNR)

        # Skriver FILLER (170 spaces)
        if not fOut.tell() == 58:
            fOut.seek(58)
        fOut.write(self.a214_FILLER)

        # Skriver VAERGEMAALS_TYPE, Værgemålstype = FULD
        if not fOut.tell() == 272:
            fOut.seek(272)
        self.a15_VAERGEMAALS_TYPE = cprDemoTool.fstr(wbs['AI' + str(row)].value, 15)
        fOut.write(f'{str(self.a15_VAERGEMAALS_TYPE):<15}' + '\n')

        self.cnt_Rec052 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 099 Hændelse
    def make_rec_099(self, row):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('099')

        # Skriver PNR
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_PNR = wbs['A' + str(row)].value  # get CPR Nummer
        fOut.write(f'{str(self.a10_PNR):>010}')

        # Skriver TIMESTAMPU, Ajourføringsdato+tid
        if not fOut.tell() == 13:
            fOut.seek(13)
        if wbs['AJ' + str(row)].value is not None:
            self.a12_TIMESTAMPU = wbs['AJ' + str(row)].value
        else:
            self.a12_TIMESTAMPU = datetime.now().strftime('%Y%m%d%H%M')
        fOut.write(f'{str(self.a12_TIMESTAMPU):0>12}')

        # Skriver HAENDELSE, Hændelse
        if not fOut.tell() == 25:
            fOut.seek(25)
        self.a03_HAENDELSE = cprDemoTool.fstr(wbs['AK' + str(row)].value, 3)
        fOut.write(f'{str(self.a03_HAENDELSE):>3}')

        # Afledtmarkering:
        # Feltet oplyser, hvilken tilknytning den ajourførte person har til hændelsens "hovedperson".
        # Blank = Hændelsens "hovedperson"
        # ÆG = Ægtefælle
        # PA = Partner
        # TÆ = Tidligere ægtefælle
        # TP = Tidligere partner
        # NÆ = Ny ægtefælle
        # NP = Ny partner
        # HÆ = Historisk ægtefælle
        # HP = Historisk partner
        # FA = Far
        # TF = Tidligere far
        # NF = Ny far
        # MO = Mor
        # TM = Tidligere mor
        # NM = Ny mor
        # BØ = Børn
        # TB = Tidligere børn
        # NB = Nye børn
        # PT = Tidligere gældende personnummer
        # PN = Nyt gældende personnummer
        # PH = Historisk personnummer
        # NV = Ny værge
        # VS = Værge slettet

        # Skriver AFLEDTMRK, Afledtmarkering
        if not fOut.tell() == 28:
            fOut.seek(28)
        self.a02_AFLEDTMRK = cprDemoTool.fstr(wbs['AL' + str(row)].value, 2)
        fOut.write(f'{str(self.a02_AFLEDTMRK):>2}')

        # Skriver NGLKONST, Nøglekonstant
        if not fOut.tell() == 30:
            fOut.seek(30)
        fOut.write(f'{str(self.a15_NGLKONST):>15}' + '\n')

        self.cnt_Rec099 += 1
        self.cnt_Set_Rec += 1

        return 0

    # 999 Slutrecord
    def make_rec_999(self):

        # Skriver RECORDTYPE
        if not fOut.tell() == 0:
            fOut.seek(0)
        fOut.write('999')

        # Skriver SORTFELT-10
        if not fOut.tell() == 3:
            fOut.seek(3)
        self.a10_SORTFELT_10 = '9999999999'
        fOut.write(f'{self.a10_SORTFELT_10:>10}')

        # Skriver TAELLER
        if not fOut.tell() == 13:
            fOut.seek(13)
        fOut.write(f'{self.cnt_Set_Rec:>08}\n')

        self.cnt_Rec999 += 1
        self.cnt_Tot_Rec = self.cnt_Set_Rec + 2

        return 0


if __name__ == '__main__':
    # Initialization
    cprDemoTool = CprDemoTool()

    # Indlæs Kørsels Data
    cprDemoTool.get_task_data()

    # Open Output file for writing with characterset ISO-8859-1
    filename = "D" + cprDemoTool.a08_PRODDTO + ".O005049"
    fOut = codecs.open(filename, "a", "ISO-8859-1")

    # Open Input file for reading CPR Related Values
    wb = load_workbook('CprChanges.xlsx')

    # Change Sheet for reading Related Values
    wbs = wb['Persondata']
    row_count = wbs.max_row
    column_count = wbs.max_column
    print("Persondata - row_count = ", row_count - 1)
    print("Persondata - column_count = ", column_count)

    # Create 000 Startrecord
    cprDemoTool.make_rec_000()

    for i in range(2, row_count + 1):
        # Create 001 Personoplysninger
        cprDemoTool.make_rec_001(i)

        # Create 002 Aktuelle adresseoplysninger
        cprDemoTool.make_rec_002(i)

        # Create 003 Adrnvn og klarskriftadresse
        cprDemoTool.make_rec_003(i)

        # Create 004 Beskyttelse
        if wbs['U' + str(i)].value:
            cprDemoTool.make_rec_004(i)

        # Create 008 Aktuelle navneoplysninger
        cprDemoTool.make_rec_008(i)

        # Create 009 Fødselsregistreringsoplysninger
        cprDemoTool.make_rec_009(i)

        # Create 010 Aktuelt statsborgerskab
        cprDemoTool.make_rec_010(i)

        # Create 017 Umyndiggørelse og værge
        # cprDemoTool.make_rec_017(i)

        # Create 052 Umyndiggørelse og værge med værgemåltype
        if wbs['AD' + str(i)].value:
            cprDemoTool.make_rec_052(i)

        # Create 099 Hændelse
        cprDemoTool.make_rec_099(i)

        # Count another Change Set done
        cprDemoTool.cnt_Tot_Set += 1

    # Create 999 Slutrecord
    cprDemoTool.make_rec_999()

    # Housekeeping
    fOut.close()
    wb.close()

    print("Der er dannet:  ", cprDemoTool.cnt_Tot_Set, " Ændringsmeddelelser fra CPR Demo Registret")
    print("   fordelt på : ", cprDemoTool.cnt_Tot_Rec, " Recoords")
    print("Der er dannet: ", cprDemoTool.cnt_Rec000, " Recoordtype 000 Startrecord")
    print("Der er dannet: ", cprDemoTool.cnt_Rec001, " Recoordtype 001 Personoplysninger")
    print("Der er dannet: ", cprDemoTool.cnt_Rec002, " Recoordtype 002 Aktuelle adresseoplysninger")
    print("Der er dannet: ", cprDemoTool.cnt_Rec003, " Recoordtype 003 Adrnvn og klarskriftadresse")
    print("Der er dannet: ", cprDemoTool.cnt_Rec004, " Recoordtype 004 Beskyttelse")
    print("Der er dannet: ", cprDemoTool.cnt_Rec008, " Recoordtype 008 Aktuelle navneoplysninger")
    print("Der er dannet: ", cprDemoTool.cnt_Rec009, " Recoordtype 009 Fødselsregistreringsoplysninger")
    print("Der er dannet: ", cprDemoTool.cnt_Rec010, " Recoordtype 010 Aktuelt statsborgerskab")
    # print("Der er dannet: ", cprDemoTool.cnt_Rec017, " Recoordtype 017 Umyndiggørelse og værge")
    print("Der er dannet: ", cprDemoTool.cnt_Rec052, " Recoordtype 052 Umyndiggørelse og værge med værgemåltype")
    print("Der er dannet: ", cprDemoTool.cnt_Rec099, " Recoordtype 099 Hændelse")
    print("Der er dannet: ", cprDemoTool.cnt_Rec999, " Recoordtype 999 Slutrecord")
